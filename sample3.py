import openpyxl as px
from openpyxl.chart import BarChart, Reference

wb = px.load_workbook('transaction.xlsx')
sheet = wb['Sheet1']


for row in range(6, sheet.max_row + 1):
    cell = sheet.cell(row, 6)
    corrected = sheet.cell(row, 7)
    corrected.value = cell.value * 0.90

val = Reference(sheet, min_row=6, max_row=sheet.max_row, min_col=7, max_col=7)

chart = BarChart()
chart.add_data(val)
sheet.add_chart(chart, 'd12')

wb.save('transactions2.xlsx')